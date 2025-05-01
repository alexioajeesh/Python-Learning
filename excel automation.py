import openpyxl as xl
wb=xl.load_workbook('transactions.xlsx')
sheet=wb['Sheet1']
cel=sheet['a1']
#or
cel=sheet.cell(1,1)
print(sheet.max_row)

for row in range(1,sheet.max_row+1):
    x=""
    for col in range(1,sheet.max_column+1):
       x=f" {x}   {sheet.cell(row,col).value}"
    print(x)

